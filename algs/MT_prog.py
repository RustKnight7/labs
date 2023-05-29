from datetime import datetime
import MetaTrader5 as mt5
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import time
import os

# выведем данные о пакете MetaTrader5
print("MetaTrader5 package author: ",mt5.__author__)
print("MetaTrader5 package version: ",mt5.__version__)

#создадим флаг итераций
it_flag = 1
# импортируем модуль pandas для вывода полученных данных в табличной форме
import pandas as pd
# импортируем модуль pytz для работы с таймзоной
import pytz


#Узнаем, какой тикер и код нужно отслеживать
tck = input('Введите требуемый тикер (код ISIN): ')

# установим подключение к терминалу MetaTrader 5
if not mt5.initialize():
    print("initialize() failed, error code =",mt5.last_error())
    quit()

#Начинаем цикл анализа данных
try:
    while True:
        utc_time = datetime.now(pytz.utc)
        actual_time_stamp = utc_time.timestamp()
        rates = mt5.copy_rates_from(tck.upper(), mt5.TIMEFRAME_M1, actual_time_stamp, 1)

        # создадим из полученных данных DataFrame
        rates_frame = pd.DataFrame(rates)
        # сконвертируем время в виде секунд в формат datetime
        rates_frame['time']=pd.to_datetime(rates_frame['time'], unit='s')
        rates_frame = rates_frame.rename(columns = {'time' : 'UTC Time', 'open' : 'Open', 'high' : 'High', 'low' : 'Low', 'close' : 'Close', 'tick_volume' : 'Tick volume', 'spread' : 'Spread', 'real_volume' : 'Real volume'})
        print(rates_frame)

        if it_flag == 1:
            wb = Workbook()
            ws = wb.active
            for r in dataframe_to_rows(rates_frame, index=False, header=True):
                ws.append(r)
            wb.save(str(tck.upper() + ".xlsx"))
        else:
            sb = load_workbook(filename=str(tck.upper() + ".xlsx"))
            ss = sb.active
            for r in dataframe_to_rows(rates_frame, index=False, header=False):
                ss.append(r)
            sb.save(str(tck.upper() + ".xlsx"))
        it_flag += 1
        for i in range(120):
            time.sleep(0.5)
except KeyboardInterrupt:
    print('Выполнение программы завершено')
    print('Файл сохранен: ', os.getcwd() + '\\' + str(tck.upper() + ".xlsx"))

# завершим подключение к терминалу MetaTrader 5
mt5.shutdown()
